"""Sync the xlsx from football-data.org (group stage) and ESPN (knockout rounds).

Does four things in one pass:
  1. Inserts any missing Group Stage MD3 rows.
  2. Backfills the Score column (col G) for finished Group Stage matches.
  3. Rebuilds ALL knockout rows from ESPN with correct dates, times, and venues.
  4. Updates scores.json so generate_ics.py shows FT scores for completed knockouts.

Both ESPN and football-data.org are used; scores are cross-checked between them.

Run after each match day or whenever the bracket advances:
    python3 update_xlsx.py
    ./update.sh          # regenerate the .ics with updated names + scores
"""

from __future__ import annotations

import hashlib
import json
import subprocess
import urllib.request
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import requests

HERE   = Path(__file__).resolve().parent
XLSX   = HERE / "2026_FIFA_World_Cup_Schedule.xlsx"
SCORES = HERE / "scores.json"
PT     = ZoneInfo("America/Los_Angeles")
YEAR   = 2026
SCORE_COL = 7

GROUP_STAGE     = "Group Stage"
KNOCKOUT_STAGES = {"Round of 32", "Round of 16", "Quarter-final",
                   "Semi-final", "Third Place", "Final"}

def stage_from_date(d: datetime) -> str | None:
    m, day = d.month, d.day
    if m == 6 and day >= 28:       return "Round of 32"
    if m == 7 and day <= 3:        return "Round of 32"
    if m == 7 and 4  <= day <= 7:  return "Round of 16"
    if m == 7 and 9  <= day <= 11: return "Quarter-final"
    if m == 7 and 14 <= day <= 15: return "Semi-final"
    if m == 7 and day == 18:       return "Third Place"
    if m == 7 and day == 19:       return "Final"
    return None

ESPN_VENUE_MAP: dict[str, tuple[str, str]] = {
    "SoFi Stadium":                    ("SoFi Stadium",          "USA"),
    "Gillette Stadium":                ("Boston Stadium",         "USA"),
    "NRG Stadium":                     ("Houston Stadium",        "USA"),
    "MetLife Stadium":                 ("MetLife Stadium",        "USA"),
    "AT&T Stadium":                    ("Dallas Stadium",         "USA"),
    "Mercedes-Benz Stadium":           ("Atlanta Stadium",        "USA"),
    "Lumen Field":                     ("Seattle Stadium",        "USA"),
    "Hard Rock Stadium":               ("Miami Stadium",          "USA"),
    "GEHA Field at Arrowhead Stadium": ("Kansas City Stadium",    "USA"),
    "BMO Field":                       ("Toronto Stadium",        "Canada"),
    "BC Place":                        ("BC Place Vancouver",     "Canada"),
    "Lincoln Financial Field":         ("Philadelphia Stadium",   "USA"),
    "Estadio BBVA":                    ("Monterrey Stadium",      "Mexico"),
    "Estadio Banorte":                 ("Guadalajara Stadium",    "Mexico"),
    "Estadio Akron":                   ("Guadalajara Stadium",    "Mexico"),
    "Levi's Stadium":                  ("San Jose Stadium",       "USA"),
    "Rose Bowl Stadium":               ("Los Angeles Stadium",    "USA"),
}

ESPN_NAME_MAP = {"United States": "USA"}
_TBD_KEYWORDS = ("Winner", "Loser", "Place", "Group", "Round",
                 "Quarterfinal", "Semifinal", "Playoff")

def norm_espn_team(name: str) -> str:
    if any(kw in name for kw in _TBD_KEYWORDS):
        return "TBD"
    return ESPN_NAME_MAP.get(name, name)

FD_STAGE_MAP = {
    "GROUP_STAGE":    GROUP_STAGE,
    "LAST_32":        "Round of 32",
    "LAST_16":        "Round of 16",
    "QUARTER_FINALS": "Quarter-final",
    "SEMI_FINALS":    "Semi-final",
    "THIRD_PLACE":    "Third Place",
    "FINAL":          "Final",
}

FD_NAME_MAP = {
    "Bosnia-Herzegovina": "Bosnia",
    "United States":      "USA",
    "Cape Verde Islands": "Cabo Verde",
    "Turkey":             "Turkiye",
}

def get_fd_token() -> str:
    r = subprocess.run(
        ["security", "find-generic-password",
         "-a", "fifa2026", "-s", "football-data-api", "-w"],
        capture_output=True, text=True,
    )
    return r.stdout.strip()

def fd_norm(name: str | None) -> str:
    if not name: return "TBD"
    return FD_NAME_MAP.get(name, name)

def fmt_date(dt: datetime) -> str:
    return dt.strftime("%B ") + str(dt.day)

def make_uid(stage: str, dt_naive: datetime, stadium: str) -> str:
    key = f"{stage}|{dt_naive.isoformat()}|{stadium}"
    return hashlib.sha1(key.encode()).hexdigest()[:16] + "@fifa-world-cup-2026"

def scrub(wb: openpyxl.Workbook) -> None:
    p = wb.properties
    for f in ("creator", "lastModifiedBy", "title", "subject",
              "description", "keywords", "category", "contentStatus"):
        setattr(p, f, "")

ESPN_URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"
KNOCKOUT_DATES = [
    "20260628","20260629","20260630","20260701","20260702","20260703",
    "20260704","20260705","20260706","20260707","20260709","20260710",
    "20260711","20260714","20260715","20260718","20260719",
]

def fetch_espn_knockout() -> list[dict]:
    fixtures = []
    for date in KNOCKOUT_DATES:
        req = urllib.request.Request(
            f"{ESPN_URL}?dates={date}",
            headers={"User-Agent": "Mozilla/5.0"},
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        for ev in data.get("events", []):
            comp    = ev["competitions"][0]
            status  = comp["status"]["type"]
            by_side = {c["homeAway"]: c for c in comp["competitors"]}
            venue_raw = comp.get("venue", {}).get("fullName", "")
            start_pt  = datetime.fromisoformat(
                ev["date"].replace("Z", "+00:00")
            ).astimezone(PT)
            stage = stage_from_date(start_pt)
            if not stage:
                continue
            stadium, country = ESPN_VENUE_MAP.get(venue_raw, (venue_raw, ""))
            home_c = by_side.get("home", {})
            away_c = by_side.get("away", {})
            home = norm_espn_team(home_c.get("team", {}).get("displayName", "TBD"))
            away = norm_espn_team(away_c.get("team", {}).get("displayName", "TBD"))
            score = ""
            if status.get("state") == "post":
                score = f"{home_c.get('score', 0)}:{away_c.get('score', 0)}"
            fixtures.append({
                "stage":    stage,
                "date":     fmt_date(start_pt),
                "time":     start_pt.strftime("%H:%M"),
                "home":     home,
                "away":     away,
                "stadium":  stadium,
                "country":  country,
                "score":    score,
                "dt_naive": start_pt.replace(tzinfo=None),
            })
    fixtures.sort(key=lambda x: x["dt_naive"])
    return fixtures


def main() -> None:
    token = get_fd_token()
    if not token:
        raise SystemExit(
            "No football-data.org token in Keychain.\n"
            "Store it: security add-generic-password "
            "-a fifa2026 -s football-data-api -w YOUR_TOKEN"
        )
    print("Fetching from football-data.org...")
    resp = requests.get(
        "https://api.football-data.org/v4/competitions/WC/matches",
        headers={"X-Auth-Token": token}, timeout=15,
    )
    resp.raise_for_status()
    raw_fd = resp.json()["matches"]
    print(f"  {len(raw_fd)} matches\n")

    fd_by_teams: dict = {}
    fd_knockout: dict = {}
    for m in raw_fd:
        home = fd_norm(m["homeTeam"].get("name"))
        away = fd_norm(m["awayTeam"].get("name"))
        if home == "TBD" or away == "TBD":
            continue
        dt_pt = datetime.fromisoformat(
            m["utcDate"].replace("Z", "+00:00")
        ).astimezone(PT)
        ft    = (m.get("score") or {}).get("fullTime") or {}
        score = (f"{ft['home']}:{ft['away']}"
                 if ft.get("home") is not None else "")
        entry = {
            "home": home, "away": away,
            "date": fmt_date(dt_pt), "time": dt_pt.strftime("%H:%M"),
            "matchday": m.get("matchday"), "status": m.get("status", ""),
            "score": score, "venue": m.get("venue") or "",
        }
        key = frozenset({home.lower(), away.lower()})
        fd_by_teams[key] = entry
        if FD_STAGE_MAP.get(m.get("stage", ""), "") in KNOCKOUT_STAGES:
            fd_knockout[key] = {"score": score, "status": m.get("status", "")}

    print("Fetching knockout schedule from ESPN...")
    knockout_fixtures = fetch_espn_knockout()
    print(f"  {len(knockout_fixtures)} fixtures\n")

    print("Cross-checking scores (ESPN vs football-data.org)...")
    mismatches = 0
    for fix in knockout_fixtures:
        if not fix["score"] or "TBD" in fix["home"] or "TBD" in fix["away"]:
            continue
        key = frozenset({fix["home"].lower(), fix["away"].lower()})
        fd  = fd_knockout.get(key)
        if fd and fd["score"] and fd["score"] != fix["score"]:
            print(f"  WARNING {fix['home']} vs {fix['away']}: "
                  f"ESPN={fix['score']} fd.org={fd['score']} -> using fd.org")
            fix["score"] = fd["score"]
            mismatches += 1
    print(f"  {mismatches} corrections\n")

    wb  = openpyxl.load_workbook(XLSX)
    ws  = wb["Schedule"]

    first_knockout_row: int | None = None
    existing_gs_teams: set = set()
    for row in ws.iter_rows(min_row=2):
        stage     = str(row[0].value or "").strip()
        match_str = str(row[3].value or "").strip()
        if stage == GROUP_STAGE and " vs " in match_str and "TBD" not in match_str:
            h, a = match_str.split(" vs ", 1)
            existing_gs_teams.add(frozenset({h.strip().lower(), a.strip().lower()}))
        if stage in KNOCKOUT_STAGES and first_knockout_row is None:
            first_knockout_row = row[0].row

    print(f"Group stage games : {len(existing_gs_teams)}")
    print(f"First knockout row: {first_knockout_row}\n")

    print("Syncing group stage dates, times and scores...")
    gs_scores = gs_times = 0
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip() != GROUP_STAGE:
            continue
        match_str = str(row[3].value or "").strip()
        if " vs " not in match_str:
            continue
        h, a  = match_str.split(" vs ", 1)
        entry = fd_by_teams.get(frozenset({h.strip().lower(), a.strip().lower()}))
        if not entry:
            continue
        if row[1].value != entry["date"] or row[2].value != entry["time"]:
            print(f"  [time]  {match_str}: {row[1].value} {row[2].value} -> {entry['date']} {entry['time']}")
            row[1].value = entry["date"]
            row[2].value = entry["time"]
            gs_times += 1
        if not row[SCORE_COL - 1].value and entry["status"] == "FINISHED" and entry["score"]:
            row[SCORE_COL - 1].value = entry["score"]
            gs_scores += 1
            print(f"  [score] {match_str} -> {entry['score']}")
    print(f"  {gs_times} times fixed, {gs_scores} scores filled\n")

    md3_missing = sorted(
        [e for k, e in fd_by_teams.items()
         if e["matchday"] == 3 and k not in existing_gs_teams],
        key=lambda x: datetime.strptime(
            f"{x['date']} {YEAR} {x['time']}", "%B %d %Y %H:%M"),
    )
    if md3_missing and first_knockout_row:
        print(f"Inserting {len(md3_missing)} missing MD3 rows...")
        ws.insert_rows(first_knockout_row, len(md3_missing))
        for i, m in enumerate(md3_missing):
            r = first_knockout_row + i
            ws.cell(row=r, column=1).value = GROUP_STAGE
            ws.cell(row=r, column=2).value = m["date"]
            ws.cell(row=r, column=3).value = m["time"]
            ws.cell(row=r, column=4).value = f"{m['home']} vs {m['away']}"
            ws.cell(row=r, column=5).value = m["venue"]
            ws.cell(row=r, column=6).value = ""
            if m["score"]:
                ws.cell(row=r, column=SCORE_COL).value = m["score"]
            print(f"  Row {r}: {m['date']} {m['time']}  "
                  f"{m['home']} vs {m['away']}"
                  + (f"  [{m['score']}]" if m["score"] else ""))
        first_knockout_row += len(md3_missing)
    else:
        print("No missing MD3 rows.\n")

    if first_knockout_row is None:
        first_knockout_row = ws.max_row + 1
    else:
        count = ws.max_row - first_knockout_row + 1
        print(f"Deleting {count} old knockout rows...")
        ws.delete_rows(first_knockout_row, count)

    print(f"Inserting {len(knockout_fixtures)} knockout rows from ESPN...")
    ws.insert_rows(first_knockout_row, len(knockout_fixtures))
    scores_json     = json.loads(SCORES.read_text()) if SCORES.exists() else {}
    new_scores_json = dict(scores_json)

    for i, fix in enumerate(knockout_fixtures):
        r         = first_knockout_row + i
        match_str = f"{fix['home']} vs {fix['away']}"
        ws.cell(row=r, column=1).value = fix["stage"]
        ws.cell(row=r, column=2).value = fix["date"]
        ws.cell(row=r, column=3).value = fix["time"]
        ws.cell(row=r, column=4).value = match_str
        ws.cell(row=r, column=5).value = fix["stadium"]
        ws.cell(row=r, column=6).value = fix["country"]
        if fix["score"]:
            ws.cell(row=r, column=SCORE_COL).value = fix["score"]
        if fix["score"] and "TBD" not in match_str:
            uid = make_uid(fix["stage"], fix["dt_naive"], fix["stadium"])
            new_scores_json[uid] = fix["score"]
        print(f"  Row {r}: [{fix['stage']}] {fix['date']} {fix['time']}  "
              f"{match_str:<35} {fix['stadium']}"
              + (f"  [{fix['score']}]" if fix["score"] else ""))

    scrub(wb)
    wb.save(XLSX)
    print(f"\nSaved {XLSX.name}")

    if new_scores_json != scores_json:
        added = len(new_scores_json) - len(scores_json)
        SCORES.write_text(json.dumps(new_scores_json, indent=2, sort_keys=True))
        print(f"Updated scores.json (+{added} knockout scores)")

    print("\nNext: ./update.sh to regenerate the .ics")


if __name__ == "__main__":
    main()
