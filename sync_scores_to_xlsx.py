"""Sync final scores from scores.json into the Score column of the xlsx.

Run manually after matches to keep the spreadsheet up to date:
    python3 sync_scores_to_xlsx.py

Never called by GitHub Actions — xlsx writes stay human-controlled.
"""

from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SOURCE_XLSX = HERE / "2026_FIFA_World_Cup_Schedule.xlsx"
SCORES_FILE = HERE / "scores.json"

YEAR = 2026
SCORE_COL = 7   # Column G


def make_uid(stage: str, dt_naive: datetime, stadium: str) -> str:
    key = f"{stage}|{dt_naive.isoformat()}|{stadium}"
    return hashlib.sha1(key.encode()).hexdigest()[:16] + "@fifa-world-cup-2026"


def scrub_metadata(wb: openpyxl.Workbook) -> None:
    p = wb.properties
    for field in ("creator", "lastModifiedBy", "title", "subject",
                  "description", "keywords", "category", "contentStatus"):
        setattr(p, field, "")


def main() -> None:
    if not SCORES_FILE.exists() or not json.loads(SCORES_FILE.read_text()):
        print("scores.json is empty — no scores to sync yet.")
        return

    scores = json.loads(SCORES_FILE.read_text())

    wb = openpyxl.load_workbook(SOURCE_XLSX)
    ws = wb["Schedule"]

    # Add Score header if not already present
    if ws.cell(row=1, column=SCORE_COL).value != "Score":
        ws.cell(row=1, column=SCORE_COL).value = "Score"

    updated = already_set = blank = skipped = 0

    for i in range(2, ws.max_row + 1):
        stage   = (ws.cell(row=i, column=1).value or "").strip()
        date_v  = ws.cell(row=i, column=2).value
        time_v  = ws.cell(row=i, column=3).value
        stadium = (ws.cell(row=i, column=5).value or "").strip()

        if not date_v or not time_v:
            continue

        try:
            dt = datetime.strptime(f"{date_v} {YEAR} {time_v}", "%B %d %Y %H:%M")
        except ValueError:
            skipped += 1
            continue

        uid = make_uid(stage, dt, stadium)
        score = scores.get(uid)
        cell = ws.cell(row=i, column=SCORE_COL)

        if score:
            if cell.value == score:
                already_set += 1
            else:
                cell.value = score
                updated += 1
        else:
            blank += 1

    scrub_metadata(wb)
    wb.save(SOURCE_XLSX)

    print(
        f"Synced {SOURCE_XLSX.name}: "
        f"{updated} written, {already_set} already correct, "
        f"{blank} still pending, {skipped} skipped"
    )


if __name__ == "__main__":
    main()
